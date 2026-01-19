import json
import base64
import os
import boto3
import openpyxl
import re
from io import BytesIO
import uuid

def handler(event: dict, context) -> dict:
    '''Загрузка Excel-файла с каталогом и обновление базы товаров'''
    
    method = event.get('httpMethod', 'POST')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        filename = body.get('filename', 'catalog.xls')
        base64_content = body.get('content', '')
        
        if not base64_content:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({
                    'success': False,
                    'error': 'Файл не передан'
                }),
                'isBase64Encoded': False
            }
        
        # Декодируем base64
        file_data = base64.b64decode(base64_content)
        
        # Загружаем в S3
        s3 = boto3.client('s3',
            endpoint_url='https://bucket.poehali.dev',
            aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
            aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY']
        )
        
        s3.put_object(
            Bucket='files',
            Key=filename,
            Body=file_data,
            ContentType='application/vnd.ms-excel'
        )
        
        # Парсим Excel и извлекаем изображения
        workbook = openpyxl.load_workbook(BytesIO(file_data))
        products_count = 0
        images_uploaded = 0
        
        for sheet in workbook.worksheets:
            sheet_name = sheet.title
            
            if 'оглавление' in sheet_name.lower():
                continue
            
            # Извлекаем изображения из листа
            if hasattr(sheet, '_images'):
                for image in sheet._images:
                    try:
                        img_data = image._data()
                        img_ext = image.format.lower()
                        img_filename = f"catalog-images/{uuid.uuid4()}.{img_ext}"
                        
                        # Загружаем изображение в S3
                        s3.put_object(
                            Bucket='files',
                            Key=img_filename,
                            Body=img_data,
                            ContentType=f'image/{img_ext}'
                        )
                        images_uploaded += 1
                    except Exception as img_error:
                        print(f'Error uploading image: {img_error}')
            
            # Ищем заголовки
            header_row_idx = None
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
                row_values = [str(val) if val else '' for val in row]
                if any('артикул' in val.lower() for val in row_values):
                    header_row_idx = row_idx
                    break
            
            if header_row_idx is None:
                continue
            
            # Считаем товары
            for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if not any(str(val).strip() if val else '' for val in row):
                    continue
                
                name_and_code = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                
                if not name_and_code:
                    continue
                
                if any(word in name_and_code.lower() for word in ['workout, комплексы', 'workout, снаряды', 'тренажеры']):
                    continue
                
                article_match = re.search(r'([А-Яа-яA-Za-z]{2,4}-\d+)', name_and_code)
                if article_match or name_and_code:
                    products_count += 1
        
        cdn_url = f"https://cdn.poehali.dev/projects/{os.environ['AWS_ACCESS_KEY_ID']}/bucket/{filename}"
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'success': True,
                'productsCount': products_count,
                'imagesUploaded': images_uploaded,
                'fileUrl': cdn_url,
                'message': f'Файл загружен успешно. Найдено товаров: {products_count}, изображений: {images_uploaded}'
            }, ensure_ascii=False),
            'isBase64Encoded': False
        }
        
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'success': False,
                'error': str(e)
            }),
            'isBase64Encoded': False
        }