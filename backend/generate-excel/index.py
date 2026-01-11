import json
import io
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
import urllib.request

def handler(event, context):
    """Генерация Excel файла с коммерческим предложением"""
    
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type'
            },
            'body': ''
        }
    
    try:
        body = json.loads(event.get('body', '{}'))
        products = body.get('products', [])
        delivery_cost = body.get('deliveryCost', 0)
        image_width = body.get('imageColumnWidth', 26)
        image_height = body.get('imageRowHeight', 99)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Коммерческое предложение"
        
        # Настройка колонок
        ws.column_dimensions['A'].width = image_width
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        # Заголовок
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Изображение', 'Артикул', 'Наименование', 'Цена, ₽', 'Кол-во', 'Сумма, ₽']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[1].height = 30
        
        # Заполнение данных
        current_row = 2
        total_sum = 0
        
        for product in products:
            ws.row_dimensions[current_row].height = image_height
            
            # Изображение
            if product.get('image') and product['image'].startswith('http'):
                try:
                    req = urllib.request.Request(product['image'], headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req, timeout=10) as response:
                        img_data = response.read()
                        img = XLImage(io.BytesIO(img_data))
                        img.width = int(image_width * 7)
                        img.height = int(image_height * 1.3)
                        ws.add_image(img, f'A{current_row}')
                except Exception as e:
                    print(f'Failed to load image: {e}')
            
            # Данные
            price = int(product['price'].replace(' ', '')) if isinstance(product['price'], str) else product['price']
            quantity = product['quantity']
            sum_price = price * quantity
            total_sum += sum_price
            
            ws.cell(row=current_row, column=2, value=product.get('article', '')).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=3, value=product.get('name', '')).alignment = Alignment(vertical='center', wrap_text=True)
            ws.cell(row=current_row, column=4, value=price).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=current_row, column=5, value=quantity).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=6, value=sum_price).alignment = Alignment(horizontal='right', vertical='center')
            
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
        
        # Доставка
        if delivery_cost > 0:
            ws.cell(row=current_row, column=5, value='Доставка:').font = Font(bold=True)
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=current_row, column=6, value=delivery_cost).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=current_row, column=6).border = border
            total_sum += delivery_cost
            current_row += 1
        
        # Итого
        ws.cell(row=current_row, column=5, value='ИТОГО:').font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=current_row, column=6, value=total_sum).font = Font(bold=True, size=12, color="4472C4")
        ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=current_row, column=6).border = border
        
        # Сохранение в память
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename="commercial_offer.xlsx"',
                'Access-Control-Allow-Origin': '*'
            },
            'body': base64.b64encode(excel_file.getvalue()).decode('utf-8'),
            'isBase64Encoded': True
        }
        
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': str(e)})
        }
